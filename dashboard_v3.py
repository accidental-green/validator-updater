import re
import requests
import openpyxl
import os
from datetime import datetime
from web3 import Web3
from web3.beacon import Beacon
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Define beacon and Web3 localhost
w3 = Web3(Web3.HTTPProvider('http://localhost:8545'))
beacon = Beacon("http://localhost:5052")

# Define the URL for data endpoints
url = "http://localhost:5052/eth/v1/beacon/headers"
url2 = "http://localhost:5052/eth/v1/node/peer_count"
url3 = "http://localhost:5052/eth/v1/node/syncing"
url4 = "http://localhost:5052/eth/v1/node/version"
url5 = "http://localhost:5052/lighthouse/peers/connected"
url6 = 'https://beaconcha.in/api/v1/validators/queue'
url7 = 'https://api.github.com/repos/sigp/lighthouse/releases'
url8 = 'https://api.github.com/repos/ethereum/go-ethereum/releases'
url9 = "https://beaconcha.in/api/v1/epoch/latest"
url10 = "https://beaconcha.in/api/v1/slot/latest"
url11 = 'https://beaconcha.in/api/v1/epoch/finalized'

# Make a GET request to the endpoints
response = requests.get(url)
response2 = requests.get(url2)
response3 = requests.get(url3)
response4 = requests.get(url4)
response5 = requests.get(url5)
response6 = requests.get(url6)
response7 = requests.get(url7)
response8 = requests.get(url8)
epoch_data = requests.get(url9).json()["data"]
slot_data = requests.get(url10).json()["data"]
final_data = requests.get(url11).json()['data']

# Extract epoch and slot from the response data
bc_current_epoch = epoch_data["epoch"]
bc_current_slot = slot_data["slot"]
bc_final_epoch = final_data['epoch']

# Lighthouse Version
lh_v = beacon.get_version()['data']['version']
match = re.search(r'v\d+\.\d+\.\d+', lh_v)
lh_ver = match.group() if match else "Version not found"

# Geth Version
geth_v = 2
match = re.search(r'v\d+\.\d+\.\d+', geth_v)
geth_ver = match.group() if match else "Version not found"

# Latest Geth version and date
geth_releases = response8.json()
geth_latest_version = geth_releases[0]['tag_name']
geth_latest_release_date_str = geth_releases[0]['published_at']
geth_latest_release_date = datetime.strptime(geth_latest_release_date_str, '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')

# Latest Sigma Prime (Lighthouse) version and date
lighthouse_releases = response7.json()
lighthouse_latest_version = lighthouse_releases[0]['tag_name']
lighthouse_latest_release_date_str = lighthouse_releases[0]['published_at']
lighthouse_latest_release_date = datetime.strptime(lighthouse_latest_release_date_str, '%Y-%m-%dT%H:%M:%SZ').strftime('%Y-%m-%d')

# Beaconcha.in Validator Count and Queue
data = response6.json()
validator_count = data['data']['validatorscount']
beaconchain_entering = data['data']['beaconchain_entering']
beaconchain_exiting = data['data']['beaconchain_exiting']

# Peers
connected_peers = response5.json()
lh_peers = len(connected_peers)
geth_peers = len(w3.geth.admin.peers())

# Geth chain ID
chain_id = w3.eth.chainId

# Get the chain name based on the ID
if chain_id == 1:
    chain_name = 'Mainnet'
elif chain_id == 3:
    chain_name = 'Ropsten Testnet'
elif chain_id == 4:
    chain_name = 'Rinkeby Testnet'
elif chain_id == 5:
    chain_name = 'Goerli Testnet'
elif chain_id == 42:
    chain_name = 'Kovan Testnet'
else:
    chain_name = 'Unknown Chain'

# Geth sync
syncing = w3.eth.syncing
if syncing:
    geth_sync = "Syncing"
else:
    geth_sync = "Synced"

# Check if node is syncing and get current/highest block number
if w3.eth.syncing:
    current_block = w3.eth.syncing['currentBlock']
    highest_block = w3.eth.syncing['highestBlock']
else:
    current_block = w3.eth.block_number
    highest_block = w3.eth.block_number

if highest_block == 0:
    geth_p = "Starting"
else:
    geth_p = (float(current_block) / float(highest_block))

# LH sync
lh_chain_id = beacon.get_deposit_contract()['data']['chain_id']
if int(lh_chain_id) == 1:
    lh_chain_name = 'Mainnet'
else:
    lh_chain_name = 'Testnet'

lh_syncing = beacon.get_syncing()
lh_syncing2 = lh_syncing['data']['is_syncing']
lh_sync = 'Syncing' if lh_syncing2 else 'Synced'

# Get finality checkpoint, epoch, and sync status for Lighthouse - localhost
finality = beacon.get_finality_checkpoint()['data']
finalized_epoch = finality['finalized']['epoch']
head_slot = lh_syncing['data']['head_slot']
sync_dist = lh_syncing['data']['sync_distance']
epoch = int(head_slot) // 32

# Set the Excel file path, open workbook and define worksheet
excel_file_path = os.path.join(os.path.expanduser("~"), "Documents", "dashboard.xlsm")
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Geth Excel Labels
worksheet["B5"] = "Geth Version"
worksheet["B6"] = "Geth Peers"
worksheet["B7"] = "Geth Sync"
worksheet["B8"] = "Current Block"
worksheet["B9"] = "Highest Block"
worksheet["B10"] = "Geth Sync %"

# Geth Excel cells
geth_ver_cell = worksheet["C5"]
geth_peers_cell = worksheet["C6"]
geth_sync_cell = worksheet["C7"]
current_block_cell = worksheet["C8"]
highest_block_cell = worksheet["C9"]
geth_p_cell = worksheet["C10"]

# Lighthouse Excel Labels
worksheet["B12"] = "Version"
worksheet["B13"] = "Peers"
worksheet["B14"] = "Sync"
worksheet["B15"] = "Sync Distance"
worksheet["B16"] = "Current Slot"
worksheet["B17"] = "Head Slot"
worksheet["B18"] = "Finalized Epoch"

# Lighthouse Excel cells
lh_curr_slot_cell = worksheet["C16"]
lh_head_slot_cell = worksheet["C17"]
lh_syncdist_cell = worksheet["C15"]
lh_ver_cell = worksheet["C12"]
lh_peers_cell = worksheet["C13"]
lh_sync_cell = worksheet["C14"]
lh_fin_epo_cell = worksheet["C18"]

# Beaconcha.in Labels Excel
worksheet["E5"] = "Current Epoch"
worksheet["E4"] = "Current Slot"
worksheet["E6"] = "Finalized Epoch"
worksheet["E7"] = "Validator Count"
worksheet["E8"] = "Enter Queue"
worksheet["E9"] = "Exit Queue"

# Beaconcha.in Excel cells
bc_epoch_cell = worksheet["F5"]
bc_slot_cell = worksheet["F4"]
bc_finalized_cell = worksheet["F6"]
bc_vcount_cell = worksheet["F7"]
beaconchain_entering_cell = worksheet["F8"]
beaconchain_exiting_cell = worksheet["F9"]

# Releases Labels Excel
worksheet["E14"] = "Geth Version"
worksheet["E15"] = "Geth Date"
worksheet["E17"] = "LH Version"
worksheet["E18"] = "LH Date"

# Releases in Excel cells
geth_latest_version_cell = worksheet["F14"]
geth_latest_release_date_cell = worksheet["F15"]
lighthouse_latest_version_cell = worksheet["F17"]
lighthouse_latest_release_date_cell = worksheet["F18"]

# Write Geth data in Excel
#geth_ver_cell.value = geth_ver
geth_peers_cell.value = geth_peers
#geth_ver_cell.value = geth_ver
geth_sync_cell.value = geth_sync
current_block_cell.value = current_block
highest_block_cell.value = highest_block
geth_p_cell.value = geth_p

# Write LH data  in Excel
lh_curr_slot_cell.value = (int(head_slot) - int(sync_dist))
lh_head_slot_cell.value = head_slot
lh_syncdist_cell.value = sync_dist
lh_ver_cell.value = lh_ver
lh_peers_cell.value = lh_peers
lh_sync_cell.value = lh_sync
lh_fin_epo_cell.value = finalized_epoch

# Write Beaconcha.in data in Excel
bc_epoch_cell.value = bc_current_epoch
bc_slot_cell.value = bc_current_slot
bc_finalized_cell.value = bc_final_epoch
bc_vcount_cell.value = validator_count
beaconchain_entering_cell.value = beaconchain_entering
beaconchain_exiting_cell.value = beaconchain_exiting

# Write release data in Excel
geth_latest_version_cell.value = geth_latest_version
geth_latest_release_date_cell.value = geth_latest_release_date
lighthouse_latest_version_cell.value = lighthouse_latest_version
lighthouse_latest_release_date_cell.value = lighthouse_latest_release_date

date_cell = worksheet['B20']
date_cell.value = "Created on " + datetime.now().strftime("%b %d, %Y at %H:%M:%S")

# Column Widths
ws = worksheet
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 9
ws.column_dimensions['D'].width = 3
ws.column_dimensions['E'].width = 14
ws.column_dimensions['F'].width = 10

# Align Cells
for column in ['B', 'E']:
    for cell in worksheet[column]:
        cell.alignment = Alignment(horizontal='left')
        
for column in ['C', 'F']:
    for cell in worksheet[column]:
        cell.alignment = Alignment(horizontal='right')

# Merge Local Node Cell
ws.merge_cells('B2:C3')
ws['B2'] = 'Local Node'
ws['B2'].font = Font(size=11, bold=True)
ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
ws['B2'].fill = PatternFill(start_color='dee6ef', end_color='dee6ef', fill_type='solid')

# Merge Geth Cell
ws.merge_cells('B4:C4')
ws['B4'] = 'Geth - ' + chain_name
ws['B4'].font = Font(size=10, bold=True)
ws['B4'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('E13:F13')
ws['E13'] = 'Geth'
ws['E13'].font = Font(size=9, bold=True)
ws['E13'].alignment = Alignment(horizontal='center', vertical='center')

# Merge Lighthouse Cell
ws.merge_cells('B11:C11')
ws['B11'] = 'Lighthouse - ' + lh_chain_name
ws['B11'].font = Font(size=10, bold=True)
ws['B11'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('E16:F16')
ws['E16'] = 'Lighthouse'
ws['E16'].font = Font(size=9, bold=True)
ws['E16'].alignment = Alignment(horizontal='center', vertical='center')

# Merge and center E2:F3 - Beaconcha.in
ws.merge_cells('E2:F3')
ws['E2'] = 'Beaconcha.in'
ws['E2'].font = Font(size=11, bold=True)
ws['E2'].alignment = Alignment(horizontal='center', vertical='center')
ws['E2'].fill = PatternFill(start_color='dde8cb', end_color='dde8cb', fill_type='solid')

# Merge Latest Release Cell
ws.merge_cells('E11:F12')
ws['E11'] = 'Latest Releases'
ws['E11'].font = Font(size=11, bold=True)
ws['E11'].alignment = Alignment(horizontal='center', vertical='center')
ws['E11'].fill = PatternFill(start_color='ffdbb6', end_color='dee6ef', fill_type='solid')

# Define the thin border
full_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

thin_border = openpyxl.styles.Side(style='thin')
bottoml_border = openpyxl.styles.Border(bottom=thin_border, left=thin_border)
bottomr_border = openpyxl.styles.Border(bottom=thin_border, right=thin_border)
topb_border = openpyxl.styles.Border(top=thin_border, bottom=thin_border)

# Loop over the cells B5 through B18
for row in worksheet.iter_rows(min_row=5, max_row=18, min_col=2, max_col=2):
    for cell in row:
        # Create a border object with a thin left border
        border = openpyxl.styles.Border(left=thin_border)

        # Set the border for the cell
        cell.border = border

# Loop over the cells C5 through C18
for row in worksheet.iter_rows(min_row=5, max_row=18, min_col=3, max_col=3):
    for cell in row:
        # Create a border object with a thin left border
        border = openpyxl.styles.Border(right=thin_border)

        # Set the border for the cell
        cell.border = border

# Loop over the cells E5 through E18
for row in worksheet.iter_rows(min_row=4, max_row=18, min_col=5, max_col=5):
    for cell in row:
        # Create a border object with a thin left border
        border = openpyxl.styles.Border(left=thin_border)

        # Set the border for the cell
        cell.border = border

# Loop over the cells F5 through F18
for row in worksheet.iter_rows(min_row=4, max_row=18, min_col=6, max_col=6):
    for cell in row:
        # Create a border object with a thin left border
        border = openpyxl.styles.Border(right=thin_border)

        # Set the border for the cell
        cell.border = border

# Apply the thin border to cells
worksheet['B2'].border = full_border
worksheet['C2'].border = full_border
worksheet['B3'].border = full_border
worksheet['C3'].border = full_border
worksheet['B4'].border = full_border
worksheet['C4'].border = full_border
worksheet['B11'].border = full_border
worksheet['C11'].border = full_border

worksheet['E2'].border = full_border
worksheet['F2'].border = full_border
worksheet['E3'].border = full_border
worksheet['F3'].border = full_border

worksheet['E11'].border = full_border
worksheet['F11'].border = full_border
worksheet['E12'].border = full_border
worksheet['F12'].border = full_border

worksheet['E13'].border = full_border
worksheet['F13'].border = full_border
worksheet['E16'].border = full_border
worksheet['F16'].border = full_border

worksheet['B18'].border = bottoml_border
worksheet['C18'].border = bottomr_border
worksheet['E18'].border = bottoml_border
worksheet['F18'].border = bottomr_border

worksheet['E10'].border = topb_border
worksheet['F10'].border = topb_border

# Set C10 as Percentage
worksheet['C10'].number_format = '0.00%'

# set the zoom to 220%
ws.sheet_view.zoomScale = 180

# Save the workbook
workbook.save(excel_file_path)

# Call the default application for opening Excel files
os.system("xdg-open " + excel_file_path)

print("Dashboard created successfully!")
